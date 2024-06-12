{
  "cells": [
    {
      "cell_type": "markdown",
      "metadata": {
        "id": "view-in-github",
        "colab_type": "text"
      },
      "source": [
        "<a href=\"https://colab.research.google.com/github/slatej1/Everside/blob/main/Parsing%20investment%20file%20into%20everside%20dataset%20template.py\" target=\"_parent\"><img src=\"https://colab.research.google.com/assets/colab-badge.svg\" alt=\"Open In Colab\"/></a>"
      ]
    },
    {
      "cell_type": "code",
      "source": [
        "import pandas as pd\n",
        "import openpyxl"
      ],
      "metadata": {
        "id": "mgDd2JDnD9z8"
      },
      "execution_count": null,
      "outputs": []
    },
    {
      "cell_type": "code",
      "source": [
        "# Read the data from the source file\n",
        "source_file_path = 'Investment Template_Farragut v4.xlsx'\n",
        "data = pd.read_excel(source_file_path, sheet_name='eFile', skiprows = 7)  # Adjust sheet_name as needed\n",
        "\n",
        "# print(data.head())\n"
      ],
      "metadata": {
        "id": "z25oa1oiEXpe"
      },
      "execution_count": null,
      "outputs": []
    },
    {
      "cell_type": "code",
      "source": [
        "column_mapping = {\n",
        "    'Company Name': 'E',\n",
        "    'Industry': 'H',\n",
        "    'Date of Investment Memo': 'D',\n",
        "    'Revenue': ['AZ', 'BO'],\n",
        "    'Adj. EBITDA ': ['BA', 'BP'],\n",
        "    'Implied EV': ['AX', 'BM'],\n",
        "    'Description': 'DE',\n",
        "    'Sourcing': 'J',\n",
        "    'Sourcing Description': 'K',\n",
        "    'Investment': 'U',\n",
        "    'Total Investment': 'R',\n",
        "    'Senior Debt.1': ['AR', 'BD'],\n",
        "    'Sr. Subordinated debt.1': ['AT', 'BF'],\n",
        "    'Cash': 'AJ',\n",
        "    'Dividend/PIK': 'AK',\n",
        "    'Cash.1': 'AO',\n",
        "    'Dividend/PIK.1': 'AP',\n",
        "    'Ownership.1': 'AQ',\n",
        "    'Tenor': 'AN',\n",
        "    'CombinedColumn': {'sources': ['Cash', 'Dividend/PIK'], 'target': 'AL'} ,\n",
        "    'CombinedColumn': {'sources': ['Senior Debt.1', 'Sr. Subordinated debt.1'], 'target': 'P'} ,\n",
        "    'SubtractedColumn': {'sources': ['Total Investment', 'Senior Debt.1', 'Sr. Subordinated debt.1'], 'target': 'O'}\n",
        "}"
      ],
      "metadata": {
        "id": "lrcoRtypEYFm"
      },
      "execution_count": null,
      "outputs": []
    },
    {
      "cell_type": "code",
      "source": [
        "# Load template for deal cloud dataset model\n",
        "template_file_path = 'deal_cloud_dataset_template.xlsx'\n",
        "workbook = load_workbook(template_file_path)\n",
        "worksheet = workbook.active"
      ],
      "metadata": {
        "id": "Yjp90qGiEYXO"
      },
      "execution_count": null,
      "outputs": []
    },
    {
      "cell_type": "code",
      "source": [
        "def col_letter_to_num(letter):\n",
        "    num = 0\n",
        "    for char in letter:\n",
        "        num = num * 26 + (ord(char.upper()) - ord('A') + 1)\n",
        "    return num"
      ],
      "metadata": {
        "id": "W-AALftVEYqS"
      },
      "execution_count": null,
      "outputs": []
    },
    {
      "cell_type": "code",
      "source": [
        "# Set the first three columns to \"Pipeline\"\n",
        "for i in range(len(data)):\n",
        "    for col_num in range(1, 4):\n",
        "        worksheet.cell(row=start_row + i, column=col_num, value=\"Pipeline\")\n",
        "\n",
        "# Write data to the template based on the column mapping\n",
        "for i, row in data.iterrows():\n",
        "    for source_col, target_cols in column_mapping.items():\n",
        "        if isinstance(target_cols, list):\n",
        "            # Handle mapping a single source column to multiple target columns\n",
        "            for target_col in target_cols:\n",
        "                value = row[source_col]\n",
        "                col_num = col_letter_to_num(target_col)  # Convert column letter to number\n",
        "                worksheet.cell(row=start_row + i, column=col_num, value=value)\n",
        "        elif isinstance(target_cols, dict) and 'sources' in target_cols:\n",
        "            # Handle combined or subtracted columns\n",
        "            sources = target_cols['sources']\n",
        "            target_col_letter = target_cols['target']\n",
        "            if source_col.startswith('Subtracted'):\n",
        "                # Subtract the sum of specified source columns\n",
        "                value = row[sources[0]] - sum(row[source] for source in sources[1:])\n",
        "            else:\n",
        "                # Combine specified columns\n",
        "                value = sum(row[source] for source in sources)\n",
        "            col_num = col_letter_to_num(target_col_letter)  # Convert column letter to number\n",
        "            worksheet.cell(row=start_row + i, column=col_num, value=value)\n",
        "        else:\n",
        "            # Single column mapping\n",
        "            value = row[source_col]\n",
        "            col_num = col_letter_to_num(target_cols)  # Convert column letter to number\n",
        "            worksheet.cell(row=start_row + i, column=col_num, value=value)\n",
        "\n",
        "\n",
        "# Save the updated template\n",
        "output_file_path = 'output_file_farragut_v4-v7.xlsx'\n",
        "workbook.save(output_file_path)"
      ],
      "metadata": {
        "id": "fHqwYfstElmF"
      },
      "execution_count": null,
      "outputs": []
    }
  ],
  "metadata": {
    "colab": {
      "name": "Welcome To Colab",
      "toc_visible": true,
      "provenance": [],
      "include_colab_link": true
    },
    "kernelspec": {
      "display_name": "Python 3",
      "name": "python3"
    }
  },
  "nbformat": 4,
  "nbformat_minor": 0
}