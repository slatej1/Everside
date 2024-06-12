import pandas as pd
import openpyxl

# Read the data from the source file
source_file_path = 'Investment Template_Farragut v4.xlsx'
data = pd.read_excel(source_file_path, sheet_name='eFile', skiprows = 7)  # Adjust sheet_name as needed

# print(data.head())

column_mapping = {
    'Company Name': 'E',
    'Industry': 'H',
    'Date of Investment Memo': 'D',
    'Revenue': ['AZ', 'BO'],
    'Adj. EBITDA ': ['BA', 'BP'],
    'Implied EV': ['AX', 'BM'],
    'Description': 'DE',
    'Sourcing': 'J',
    'Sourcing Description': 'K',
    'Investment': 'U',
    'Total Investment': 'R',
    'Senior Debt.1': ['AR', 'BD'],
    'Sr. Subordinated debt.1': ['AT', 'BF'],
    'Cash': 'AJ',
    'Dividend/PIK': 'AK',
    'Cash.1': 'AO',
    'Dividend/PIK.1': 'AP',
    'Ownership.1': 'AQ',
    'Tenor': 'AN',
    'CombinedColumn': {'sources': ['Cash', 'Dividend/PIK'], 'target': 'AL'} ,
    'CombinedColumn': {'sources': ['Senior Debt.1', 'Sr. Subordinated debt.1'], 'target': 'P'} ,
    'SubtractedColumn': {'sources': ['Total Investment', 'Senior Debt.1', 'Sr. Subordinated debt.1'], 'target': 'O'} 
}

template_file_path = 'deal_cloud_dataset_template.xlsx'
workbook = load_workbook(template_file_path)
worksheet = workbook.active

def col_letter_to_num(letter):
    num = 0
    for char in letter:
        num = num * 26 + (ord(char.upper()) - ord('A') + 1)
    return num

# Set the first three columns to "Pipeline"
for i in range(len(data)):
    for col_num in range(1, 4):
        worksheet.cell(row=start_row + i, column=col_num, value="Pipeline")

# Write data to the template based on the column mapping
for i, row in data.iterrows():
    for source_col, target_cols in column_mapping.items():
        if isinstance(target_cols, list):
            # Handle mapping a single source column to multiple target columns
            for target_col in target_cols:
                value = row[source_col]
                col_num = col_letter_to_num(target_col)  # Convert column letter to number
                worksheet.cell(row=start_row + i, column=col_num, value=value)
        elif isinstance(target_cols, dict) and 'sources' in target_cols:
            # Handle combined or subtracted columns
            sources = target_cols['sources']
            target_col_letter = target_cols['target']
            if source_col.startswith('Subtracted'):
                # Subtract the sum of specified source columns
                value = row[sources[0]] - sum(row[source] for source in sources[1:])
            else:
                # Combine specified columns
                value = sum(row[source] for source in sources)
            col_num = col_letter_to_num(target_col_letter)  # Convert column letter to number
            worksheet.cell(row=start_row + i, column=col_num, value=value)
        else:
            # Single column mapping
            value = row[source_col]
            col_num = col_letter_to_num(target_cols)  # Convert column letter to number
            worksheet.cell(row=start_row + i, column=col_num, value=value)


# Save the updated template
output_file_path = 'output_file_farragut_v4-v7.xlsx'
workbook.save(output_file_path)
